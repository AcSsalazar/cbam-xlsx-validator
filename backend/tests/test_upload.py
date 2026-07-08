"""Functional tests for ``POST /upload``.

Each test uses the ``client_with_db`` fixture which provides a TestClient
backed by a fresh file-based SQLite database with the schema already
created. After every test the schema is dropped, so tests are isolated.
"""
from __future__ import annotations

from io import BytesIO

import openpyxl
from tests._xlsx_helpers import RULES_HEADERS, RULES_ROWS


def test_upload_happy_path(client_with_db, sample_xlsx_bytes):
    response = client_with_db.post(
        "/upload",
        files={"file": ("sample.xlsx", sample_xlsx_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert response.status_code == 200
    body = response.json()
    assert body["filename"] == "sample.xlsx"
    assert body["total_rows"] == 5
    assert body["valid_rows"] == 1
    assert body["invalid_rows"] == 4
    # Errors are emitted per field-level violation, not per row.
    assert len(body["errors"]) > 4
    for error in body["errors"]:
        assert set(error.keys()) == {"row", "field", "value", "message"}


def test_upload_persists_only_valid_rows(client_with_db, sample_xlsx_bytes):
    client_with_db.post(
        "/upload",
        files={"file": ("sample.xlsx", sample_xlsx_bytes, "application/vnd.ms-excel")},
    )
    response = client_with_db.get("/records?page=1&page_size=20")
    assert response.status_code == 200
    body = response.json()
    assert body["total"] == 1
    assert len(body["items"]) == 1
    persisted = body["items"][0]
    assert persisted["row_number"] == 2
    assert persisted["payload"]["EORI Number"] == "DE123456789012345"
    assert persisted["payload"]["CN Code"] == "72071114"


def test_upload_invalid_file_extension(client_with_db):
    response = client_with_db.post(
        "/upload",
        files={"file": ("data.csv", b"a,b,c\n1,2,3", "text/csv")},
    )
    assert response.status_code == 400
    assert ".xlsx" in response.json()["detail"]


def test_upload_invalid_xlsx_bytes(client_with_db):
    response = client_with_db.post(
        "/upload",
        files={"file": ("sample.xlsx", b"this is not an xlsx", "application/octet-stream")},
    )
    assert response.status_code == 400


def test_upload_missing_template_sheet(client_with_db):
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)
    rules = workbook.create_sheet(title="Rules")
    rules.append(RULES_HEADERS)
    for row in RULES_ROWS:
        rules.append(row)
    buffer = BytesIO()
    workbook.save(buffer)

    response = client_with_db.post(
        "/upload",
        files={"file": ("bad.xlsx", buffer.getvalue(), "application/octet-stream")},
    )
    assert response.status_code == 400
    assert "Template" in response.json()["detail"]


def test_upload_missing_rules_sheet(client_with_db):
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)
    template = workbook.create_sheet(title="Template")
    template.append(["EORI Number"])
    buffer = BytesIO()
    workbook.save(buffer)

    response = client_with_db.post(
        "/upload",
        files={"file": ("bad.xlsx", buffer.getvalue(), "application/octet-stream")},
    )
    assert response.status_code == 400
    assert "rules sheet" in response.json()["detail"].lower()


def test_upload_header_mismatch(client_with_db):
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)
    template = workbook.create_sheet(title="Template")
    template.append(["EORI Number", "Unknown Column"])
    rules = workbook.create_sheet(title="Rules")
    rules.append(RULES_HEADERS)
    for row in RULES_ROWS:
        rules.append(row)
    buffer = BytesIO()
    workbook.save(buffer)

    response = client_with_db.post(
        "/upload",
        files={"file": ("bad.xlsx", buffer.getvalue(), "application/octet-stream")},
    )
    assert response.status_code == 400
    assert "Unknown Column" in response.json()["detail"]


def test_upload_accepts_sheet1_as_rules_sheet(client_with_db, sample_xlsx_bytes_sheet1):
    response = client_with_db.post(
        "/upload",
        files={"file": ("sample.xlsx", sample_xlsx_bytes_sheet1, "application/octet-stream")},
    )
    assert response.status_code == 200
    assert response.json()["valid_rows"] == 1
