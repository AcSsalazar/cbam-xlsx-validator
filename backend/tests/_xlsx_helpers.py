"""Shared helpers for building CBAM-shaped .xlsx files in tests.

Both the conftest fixtures and the per-test record builders reuse the
data and builder here so the rules-sheet structure stays in one place.
"""
from __future__ import annotations

from io import BytesIO

import openpyxl

TEMPLATE_COLUMNS = [
    "EORI Number", "Declarant Legal Name", "Declarant Address",
    "Contact Person", "Competent Authority", "CBAM Account Number",
    "Data Owner", "TARIC Code", "CN Code", "Goods Description",
    "Sector Category", "Product Type", "Import Volume",
    "Date of importation", "Country of Origin", "Customs Declaration Ref",
    "Supplier Name", "Notes / Comments",
]

RULES_HEADERS = [
    "Field", "Purpose",
    "Mandatory/optional for compliance",
    "Mandatory/optional for impact assessment",
    "Data type/format", "Validation rule", "Example value",
    "Minimum fields required for CBAM scope classification",
    "Notes/assumptions",
]

RULES_ROWS = [
    ["EORI Number", "To identify declarant, required by law", "Mandatory ", "No",
     "Text (up to 17 chars)",
     "Format: 2-letter country code + up to 15 alphanumeric chars",
     "DE123456789012345", "No", "one row per declarant"],
    ["Declarant Legal Name", "To identify declarant, required by law", "Mandatory ", "No",
     "Text", "Non-empty", "ArcelorMittal SA", "No", "one row per declarant"],
    ["Declarant Address", "To identify declarant, required by law", "Mandatory ", "No",
     "Text", "Non-empty (We can check that the address exists?)",
     "24-26 Boulevard d'Avranches, L-1160 Luxembourg ", "No", "one row per declarant"],
    ["Contact Person", "To identify declarant's responsible person, required by law",
     "Mandatory ", "No", "Text",
     "Must include name and at least one contact method",
     "John Doe, john.doe@company.com", "No", "one row per declarant"],
    ["Competent Authority", "Required by law", "Mandatory ", "No", "Text",
     "Must match a valid EU Member State competent authority", "DEHSt", "No",
     "one row per declarant"],
    ["CBAM Account Number", "Required by law", "Mandatory ", "No", "Text", "n/a",
     "CBAM-DE-2026-00142", "No", "one row per declarant"],
    ["Data Owner", "Identifies the internal person/team responsible for the data.",
     "Mandatory ", "No", "Text", "Non-empty; must be a valid internal contact",
     "Sam Smith, sam.smith@company.com", "No", None],
    ["TARIC Code", "To identify assigned code for the import", "Optional", "No",
     "Text (10 digits)", "Format: 10 digits", 7207111400, "No", None],
    ["CN Code", "Used to determine whether the product may be within CBAM scope.",
     "Mandatory ", "Yes", "Numeric (8 digits)",
     "Must be exactly 8 digits; must fall within CBAM Annex I codes",
     72071114, "Yes", "One row per one CN Code +supplier + country"],
    ["Goods Description", "Full product description matching the CN code classification",
     "Mandatory ", "Yes", "Text", "Must match CN Code classification",
     "Semi-finished iron or non-alloy steel (carbon <0.25%) with a rectangular cross-section and a thickness not exceeding 130 mm",
     "Yes", None],
    ["Sector Category", "Helps classify the product under CBAM-relevant categories.",
     "Mandatory ", "No", "Text", "Must match CN Code classification",
     "Iron and Steel", "No", None],
    ["Product Type", "Simple Goods (no precursors) or Complex Goods (contains CBAM precursors)",
     "Mandatory ", "No", "Enum", None, "Complex", "No", None],
    ["Import Volume",
     "Net mass of imported goods in metric tonnes for the reporting/analysing period period",
     "Mandatory ", "Yes", "Decimal (tonnes)", "Must be a positive number", 1250, "No", None],
    ["Date of importation", "Indicates the period the data belongs to.",
     "Mandatory ", "Yes", "Integer", None, "05.05.2026", "No",
     "Reporting period - 1 year, analysing period could be any period"],
    ["Country of Origin", "Country of production, required for CBAM scope determination.",
     "Mandatory ", "Yes", "ISO 3166-1",
     "Must match ISO 3166-1, for CBAM scope must be outside EU + exemption countries from the regulation",
     "China", "Yes", None],
    ["Customs Declaration Ref", "Reference number of the customs import declaration",
     "Optional", "No", "Text", "Non-empty", "DE/2026/MRN-ABC-123456", "No", None],
    ["Supplier Name", "Links the product to the supplier providing or producing the goods.",
     "Mandatory ", "Yes", "Text", "Non-empty", "Supplier Ch1", "No", None],
    ["Notes / Comments", "Captures additional context or known data limitations.",
     "Optional", "No", "Text", None, "MRV plan is under preparation", "No", None],
]


def build_xlsx(
    template_rows: list[list],
    rules_sheet_name: str = "Rules",
) -> bytes:
    """Build a CBAM-shaped .xlsx in memory and return its bytes."""
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)

    template = workbook.create_sheet(title="Template")
    template.append(TEMPLATE_COLUMNS)
    for row in template_rows:
        template.append(row)

    rules = workbook.create_sheet(title=rules_sheet_name)
    rules.append(RULES_HEADERS)
    for row in RULES_ROWS:
        rules.append(row)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
